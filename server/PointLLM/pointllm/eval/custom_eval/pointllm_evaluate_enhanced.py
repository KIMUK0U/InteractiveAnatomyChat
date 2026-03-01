"""
PointLLM Evaluation with Enhanced Excel Output
This version uses openpyxl for better image handling in Excel files
"""

import argparse
import os
import sys
import glob
import json
import torch
import numpy as np
import pandas as pd
from PIL import Image
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import warnings
warnings.filterwarnings('ignore')

# Import PointLLM modules
sys.path.append('PointLLM')
from pointllm.conversation import conv_templates, SeparatorStyle
from pointllm.utils import disable_torch_init
from pointllm.model import PointLLMLlamaForCausalLM
from pointllm.model.utils import KeywordsStoppingCriteria
from transformers import AutoTokenizer

def pc_normalize(pc):
    """
    Normalize a point cloud to fit within a unit sphere
    pc: Nx6 or Nx3 array (xyz + optional rgb)
    """
    xyz = pc[:, :3]
    other_features = pc[:, 3:] if pc.shape[1] > 3 else None
    
    # Center the point cloud
    centroid = np.mean(xyz, axis=0)
    xyz = xyz - centroid
    
    # Scale to unit sphere
    m = np.max(np.sqrt(np.sum(xyz**2, axis=1)))
    if m > 0:
        xyz = xyz / m
    
    # Combine normalized xyz with other features
    if other_features is not None:
        pc = np.concatenate((xyz, other_features), axis=1)
    else:
        pc = xyz
    
    return pc

def load_custom_point_cloud(npy_path, use_color=True):
    """
    Load and preprocess a custom point cloud from .npy file
    """
    print(f"Loading point cloud from: {npy_path}")
    
    # Load the point cloud
    point_cloud = np.load(npy_path)
    print(f"  Original shape: {point_cloud.shape}")
    
    # Ensure we have the right format (N, 6) for xyz+rgb
    if point_cloud.shape[1] < 6 and use_color:
        # If no color, add zeros
        zeros_rgb = np.zeros((point_cloud.shape[0], 3))
        point_cloud = np.concatenate([point_cloud[:, :3], zeros_rgb], axis=1)
        print("  No RGB values found, adding zeros")
    elif point_cloud.shape[1] >= 6:
        # Take only xyz and rgb
        point_cloud = point_cloud[:, :6]
    
    # Normalize xyz coordinates
    point_cloud = pc_normalize(point_cloud)
    
    # Ensure RGB values are in [0, 1]
    if use_color and point_cloud.shape[1] >= 6:
        rgb = point_cloud[:, 3:6]
        if np.max(rgb) > 1.0:
            point_cloud[:, 3:6] = rgb / 255.0
            print("  Normalized RGB from [0-255] to [0-1]")
    
    # Sample to 8192 points if necessary
    if point_cloud.shape[0] > 8192:
        indices = np.random.choice(point_cloud.shape[0], 8192, replace=False)
        point_cloud = point_cloud[indices]
        print(f"  Downsampled to 8192 points")
    elif point_cloud.shape[0] < 8192:
        # Repeat points if we have too few
        repeats = 8192 // point_cloud.shape[0] + 1
        point_cloud = np.tile(point_cloud, (repeats, 1))[:8192]
        print(f"  Upsampled to 8192 points")
    
    print(f"  Final shape: {point_cloud.shape}")
    
    return point_cloud

def visualize_point_cloud_multi_view(point_cloud, save_path, title="Point Cloud"):
    """
    Create and save a multi-view visualization of the point cloud
    """
    fig = plt.figure(figsize=(16, 12))
    
    # Extract xyz and rgb
    xyz = point_cloud[:, :3]
    
    # Use RGB if available, otherwise use default color
    if point_cloud.shape[1] >= 6:
        colors = point_cloud[:, 3:6]
    else:
        colors = np.ones((point_cloud.shape[0], 3)) * 0.5
    
    # Create 4 subplots for different views
    views = [
        (221, 30, 45, "3D View (30°, 45°)"),
        (222, 0, 90, "Top View"),
        (223, 0, 0, "Front View"),
        (224, 90, 0, "Side View")
    ]
    
    for subplot_id, elev, azim, view_title in views:
        ax = fig.add_subplot(subplot_id, projection='3d')
        
        # Plot points with smaller size for better visibility
        ax.scatter(xyz[:, 0], xyz[:, 1], xyz[:, 2], 
                  c=colors, s=0.5, alpha=0.6)
        
        # Set view angle
        ax.view_init(elev=elev, azim=azim)
        
        # Set labels and title
        ax.set_xlabel('X', fontsize=8)
        ax.set_ylabel('Y', fontsize=8)
        ax.set_zlabel('Z', fontsize=8)
        ax.set_title(view_title, fontsize=10)
        
        # Set equal aspect ratio
        max_range = np.array([xyz[:, 0].max()-xyz[:, 0].min(),
                              xyz[:, 1].max()-xyz[:, 1].min(),
                              xyz[:, 2].max()-xyz[:, 2].min()]).max() / 2.0
        
        mid_x = (xyz[:, 0].max()+xyz[:, 0].min()) * 0.5
        mid_y = (xyz[:, 1].max()+xyz[:, 1].min()) * 0.5
        mid_z = (xyz[:, 2].max()+xyz[:, 2].min()) * 0.5
        
        ax.set_xlim(mid_x - max_range, mid_x + max_range)
        ax.set_ylim(mid_y - max_range, mid_y + max_range)
        ax.set_zlim(mid_z - max_range, mid_z + max_range)
        
        # Remove grid for cleaner look
        ax.grid(True, alpha=0.3)
    
    # Main title
    fig.suptitle(title, fontsize=14, fontweight='bold')
    
    # Adjust layout and save
    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"  Multi-view visualization saved to: {save_path}")

def load_prompts(prompt_file):
    """
    Load prompts from a JSON file
    """
    with open(prompt_file, 'r') as f:
        prompts_data = json.load(f)
    return prompts_data

def init_model(model_path, device='cuda'):
    """
    Initialize PointLLM model
    """
    print(f"Initializing model from: {model_path}")
    
    disable_torch_init()
    
    # Load tokenizer and model
    tokenizer = AutoTokenizer.from_pretrained(model_path)
    
    # Determine torch dtype and device
    if device == 'cuda' and torch.cuda.is_available():
        torch_dtype = torch.float16
        device = 'cuda'
    elif device == 'mps' and torch.backends.mps.is_available():
        torch_dtype = torch.float32
        device = 'mps'
    else:
        torch_dtype = torch.float32
        device = 'cpu'
    
    model = PointLLMLlamaForCausalLM.from_pretrained(
        model_path, 
        low_cpu_mem_usage=True, 
        use_cache=True, 
        torch_dtype=torch_dtype
    )
    
    model = model.to(device)
    model.initialize_tokenizer_point_backbone_config_wo_embedding(tokenizer)
    model.eval()
    
    # Get point configuration
    mm_use_point_start_end = getattr(model.config, "mm_use_point_start_end", False)
    point_backbone_config = model.get_model().point_backbone_config
    
    # Setup conversation template
    conv_mode = "vicuna_v1_1"
    conv = conv_templates[conv_mode].copy()
    
    stop_str = conv.sep if conv.sep_style != SeparatorStyle.TWO else conv.sep2
    keywords = [stop_str]
    
    return model, tokenizer, point_backbone_config, keywords, mm_use_point_start_end, conv, device

def run_inference(model, tokenizer, point_cloud, prompt_text, 
                 point_backbone_config, keywords, mm_use_point_start_end, conv, device):
    """
    Run inference on a single point cloud with a given prompt
    """
    # Reset conversation
    conv.reset()
    
    # Prepare point cloud tokens
    point_token_len = point_backbone_config['point_token_len']
    default_point_patch_token = point_backbone_config['default_point_patch_token']
    default_point_start_token = point_backbone_config['default_point_start_token']
    default_point_end_token = point_backbone_config['default_point_end_token']
    
    # Prepare the question with point tokens
    if mm_use_point_start_end:
        qs = default_point_start_token + default_point_patch_token * point_token_len + \
             default_point_end_token + '\n' + prompt_text
    else:
        qs = default_point_patch_token * point_token_len + '\n' + prompt_text
    
    # Add to conversation
    conv.append_message(conv.roles[0], qs)
    conv.append_message(conv.roles[1], None)
    prompt = conv.get_prompt()
    
    # Tokenize
    inputs = tokenizer([prompt])
    input_ids = torch.as_tensor(inputs.input_ids).to(device)
    
    # Prepare point cloud tensor
    point_cloud_tensor = torch.from_numpy(point_cloud).unsqueeze(0).to(device).to(model.dtype)
    
    # Setup stopping criteria
    stopping_criteria = KeywordsStoppingCriteria(keywords, tokenizer, input_ids)
    stop_str = keywords[0]
    
    # Generate response with adjusted parameters
    with torch.inference_mode():
        output_ids = model.generate(
            input_ids,
            point_clouds=point_cloud_tensor,
            do_sample=True,
            temperature=0.7,
            top_k=50,
            max_new_tokens=512,  # Limit response length
            top_p=0.95,
            repetition_penalty=1.2,  # Increase to reduce repetition
            no_repeat_ngram_size=3,  # Prevent 3-gram repetitions
            stopping_criteria=[stopping_criteria]
        )
    
    # Decode output
    input_token_len = input_ids.shape[1]
    outputs = tokenizer.batch_decode(output_ids[:, input_token_len:], skip_special_tokens=True)[0]
    outputs = outputs.strip()
    
    if outputs.endswith(stop_str):
        outputs = outputs[:-len(stop_str)]
    
    outputs = outputs.strip()
    
    return outputs

def create_excel_with_openpyxl(results_df, image_paths, prompts_data, output_path):
    """
    Create an Excel file with embedded images using openpyxl
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"
    
    # Define styles
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    prompt_def_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, color="FFFFFF")
    
    # Get prompt keys
    prompt_keys = list(prompts_data['prompts'].keys())
    
    # Write headers
    headers = ['Object ID', 'Visualization'] + prompt_keys
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write prompt definitions in second row
    ws.cell(row=2, column=1, value="")
    ws.cell(row=2, column=2, value="")
    for col_idx, prompt_key in enumerate(prompt_keys, 3):
        cell = ws.cell(row=2, column=col_idx, 
                      value=prompts_data['prompts'][prompt_key]['definition'])
        cell.fill = prompt_def_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        if col_letter <= chr(ord('B') + len(prompt_keys)):
            ws.column_dimensions[col_letter].width = 50
    
    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 60
    
    # Write data and insert images
    for row_idx, (result, img_path) in enumerate(zip(results_df, image_paths), 3):
        # Object ID
        cell = ws.cell(row=row_idx, column=1, value=result['num_id'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set row height for image
        ws.row_dimensions[row_idx].height = 200
        
        # Insert image
        if os.path.exists(img_path):
            img = XLImage(img_path)
            # Resize image
            img.width = 280
            img.height = 200
            # Position image in cell
            img.anchor = f'B{row_idx}'
            ws.add_image(img)
        
        # Write prompt responses
        for col_idx, prompt_key in enumerate(prompt_keys, 3):
            cell = ws.cell(row=row_idx, column=col_idx, value=result.get(prompt_key, ''))
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel file with images saved to: {output_path}")

def main(args):
    # Create output directory
    os.makedirs(args.output_dir, exist_ok=True)
    
    # Load prompts
    prompts_data = load_prompts(args.prompt_file)
    prompt_keys = list(prompts_data['prompts'].keys())
    
    # Initialize model
    model, tokenizer, point_backbone_config, keywords, mm_use_point_start_end, conv, device = \
        init_model(args.model_path, args.device)
    
    # Find all .npy files
    npy_files = glob.glob(os.path.join(args.data_path, '*.npy'))
    print(f"\nFound {len(npy_files)} .npy files")
    
    # Limit number of files for testing
    if args.max_objects > 0:
        npy_files = npy_files[:args.max_objects]
        print(f"Processing first {args.max_objects} objects")
    
    # Prepare data for DataFrame
    results = []
    image_paths = []
    
    # Create visualization directory
    viz_dir = os.path.join(args.output_dir, 'visualizations')
    os.makedirs(viz_dir, exist_ok=True)
    
    # Process each point cloud
    for idx, npy_path in enumerate(npy_files):
        filename = os.path.basename(npy_path)
        object_name = os.path.splitext(filename)[0]
        
        print(f"\n[{idx+1}/{len(npy_files)}] Processing: {object_name}")
        
        # Load point cloud
        try:
            point_cloud = load_custom_point_cloud(npy_path, use_color=True)
        except Exception as e:
            print(f"  Error loading point cloud: {e}")
            continue
        
        # Create multi-view visualization
        viz_path = os.path.join(viz_dir, f"{object_name}.png")
        visualize_point_cloud_multi_view(point_cloud, viz_path, title=object_name)
        image_paths.append(viz_path)
        
        # Run inference for each prompt
        row_data = {
            'num_id': f"{idx + 1:03d}",
            'object_name': object_name
        }
        
        for prompt_key in prompt_keys:
            prompt_text = prompts_data['prompts'][prompt_key]['full_prompt']
            print(f"  Running inference for {prompt_key}...")
            
            try:
                response = run_inference(
                    model, tokenizer, point_cloud, prompt_text,
                    point_backbone_config, keywords, mm_use_point_start_end, conv, device
                )
                # Clean up response
                response = response.replace('\n\n', ' ').replace('  ', ' ').strip()
                row_data[prompt_key] = response
                print(f"    Response preview: {response[:80]}...")
            except Exception as e:
                print(f"    Error during inference: {e}")
                row_data[prompt_key] = f"Error: {str(e)}"
        
        results.append(row_data)
        
        # Save intermediate results every 5 objects
        if (idx + 1) % 5 == 0:
            temp_df = pd.DataFrame(results)
            temp_df.to_csv(os.path.join(args.output_dir, 'temp_results.csv'), index=False)
            print(f"  Saved intermediate results ({idx + 1} objects processed)")
    
    # Save final results
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Save CSV version
    results_df = pd.DataFrame(results)
    csv_path = os.path.join(args.output_dir, f'evaluation_results_{timestamp}.csv')
    results_df.to_csv(csv_path, index=False)
    print(f"\nCSV results saved to: {csv_path}")
    
    # Save Excel version with images
    excel_path = os.path.join(args.output_dir, f'evaluation_results_{timestamp}.xlsx')
    create_excel_with_openpyxl(results, image_paths, prompts_data, excel_path)
    
    # Save HTML version for easy viewing
    html_content = generate_html_report(results, image_paths, prompts_data)
    html_path = os.path.join(args.output_dir, f'evaluation_results_{timestamp}.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"HTML report saved to: {html_path}")
    
    # Save summary statistics
    summary = {
        'total_objects': len(results),
        'prompts_used': prompt_keys,
        'model_path': args.model_path,
        'device': device,
        'timestamp': timestamp
    }
    
    with open(os.path.join(args.output_dir, f'evaluation_summary_{timestamp}.json'), 'w') as f:
        json.dump(summary, f, indent=2)
    
    print("\n" + "="*50)
    print("Evaluation completed successfully!")
    print(f"Results saved to: {args.output_dir}")
    print("="*50)

def generate_html_report(results, image_paths, prompts_data):
    """
    Generate an HTML report for easy viewing of results
    """
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>PointLLM Evaluation Results</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            table { border-collapse: collapse; width: 100%; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            th { background-color: #4CAF50; color: white; }
            .prompt-def { background-color: #E8F5E9; font-style: italic; }
            .visualization { max-width: 400px; height: auto; }
            .response { white-space: pre-wrap; word-wrap: break-word; max-width: 400px; }
        </style>
    </head>
    <body>
        <h1>PointLLM Evaluation Results</h1>
        <table>
    """
    
    # Add header row
    html += "<tr><th>ID</th><th>Visualization</th>"
    for prompt_key in prompts_data['prompts'].keys():
        html += f"<th>{prompt_key}</th>"
    html += "</tr>"
    
    # Add prompt definition row
    html += "<tr><td></td><td></td>"
    for prompt_key in prompts_data['prompts'].keys():
        html += f"<td class='prompt-def'>{prompts_data['prompts'][prompt_key]['definition']}</td>"
    html += "</tr>"
    
    # Add data rows
    for result, img_path in zip(results, image_paths):
        html += f"<tr><td>{result['num_id']}</td>"
        
        # Add image
        img_rel_path = os.path.relpath(img_path, os.path.dirname(img_path).replace('visualizations', ''))
        html += f"<td><img src='{img_rel_path}' class='visualization'/></td>"
        
        # Add responses
        for prompt_key in prompts_data['prompts'].keys():
            response = result.get(prompt_key, '')
            html += f"<td class='response'>{response}</td>"
        html += "</tr>"
    
    html += """
        </table>
    </body>
    </html>
    """
    
    return html

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Evaluate PointLLM on custom point clouds")
    
    parser.add_argument("--model_path", type=str,
                       default="RunsenXu/PointLLM_7B_v1.2",
                       help="Path to PointLLM model")
    
    parser.add_argument("--data_path", type=str,
                       default="./pointllm/data/npy",
                       help="Path to directory containing .npy files")
    
    parser.add_argument("--prompt_file", type=str,
                       default="./pointllm/eval/custom_eval/prompts.json",
                       help="Path to JSON file containing prompts")
    
    parser.add_argument("--output_dir", type=str,
                       default="./pointllm/data/evaluation_results",
                       help="Directory to save results")
    
    parser.add_argument("--device", type=str, default="mps",
                       choices=["cuda", "cpu", "mps"],
                       help="Device to use for inference")
    
    parser.add_argument("--max_objects", type=int, default=-1,
                       help="Maximum number of objects to process (-1 for all)")
    
    args = parser.parse_args()
    
    main(args)
